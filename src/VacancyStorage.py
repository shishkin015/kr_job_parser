import csv
import json
from abc import ABC, abstractmethod
from json import JSONDecodeError

import openpyxl
import psycopg2

from src.vacancy import Vacancy


class VacancyStorage(ABC):

    @abstractmethod
    def load_from_file(self):
        """
        Абстрактный метод для загрузки данных о вакансиях из файла в хранилище.
        """
        pass

    @abstractmethod
    def save_to_file(self):
        """
        Абстрактный метод для сохранения данных о вакансиях из хранилища в файл.
        """
        pass

    @abstractmethod
    def add_vacancy(self, vacancy):
        """
        Абстрактный метод для добавления вакансии в хранилище.

        :param vacancy: Экземпляр класса Vacancy, который нужно добавить.
        """
        pass

    @abstractmethod
    def get_vacancies(self, criteria):
        """
        Абстрактный метод для поиска вакансий в хранилище по заданным критериям.

        :param criteria: Критерии поиска вакансий.
        :return: Список вакансий, соответствующих заданным критериям.
        """
        pass

    @abstractmethod
    def remove_vacancy(self, vacancy_id):
        """
        Абстрактный метод для удаления вакансии из хранилища по её ID.

        :param vacancy_id: ID вакансии, которую нужно удалить.
        """
        pass


class JsonVacancyStorage(VacancyStorage):

    def __init__(self, path: str, vacancies: list[dict] = None):
        self._path = path
        self.dict_vacancies = vacancies
        self.instance_vacancies = []

    def load_from_file(self):
        try:

            try:
                with open(self._path, "r", encoding="utf-8") as f:
                    json_data = json.load(f)
            except JSONDecodeError:
                print("Файл пуст")
                return

            for item in json_data:
                vacancy = Vacancy(
                    int(item.get('id_num', "Нет данных")),
                    str(item.get('vacancy_name', "Нет данных")),
                    str(item.get('company_name')),
                    str(item.get('description', "Нет данных")),
                    float(item.get('salary_from')) if item.get('salary_from') is not None else None,
                    float(item.get('salary_to')) if item.get('salary_to') is not None else None,
                    str(item.get('salary_currency')),
                    str(item.get('area', "Нет данных")),
                    str(item.get('url', "Нет данных")),
                )
                self.instance_vacancies.append(vacancy)
            self.dict_vacancies = [vacancy.to_dict() for vacancy in self.instance_vacancies]
        except FileNotFoundError:
            self.dict_vacancies = []
            raise FileNotFoundError("File not found")

    def save_to_file(self):
        with open(self._path, "w", encoding="utf-8") as f:
            json.dump(self.dict_vacancies, f, indent=4)

    def add_vacancy(self, vacancy):
        self.load_from_file()
        self.dict_vacancies.append(vacancy)
        self.save_to_file()

    def get_vacancies(self, criteria):
        self.load_from_file()
        matching_vacancies = []
        for vacancy in self.instance_vacancies:
            dict_vacancy = vacancy.to_dict()
            if criteria in dict_vacancy['vacancy_name'] or criteria in dict_vacancy['description']:
                matching_vacancies.append(vacancy)
        return matching_vacancies

    def remove_vacancy(self, vacancy_id):
        self.load_from_file()

        try:
            self.dict_vacancies = [vacancy for vacancy in self.dict_vacancies if vacancy['id_num'] != int(vacancy_id)]
            self.save_to_file()
            print("Вакансия успешно удалена!")
        except ValueError:
            print("Вакансия с указанным ID не найдена.")


class CSVVacancyStorage(VacancyStorage):

    def __init__(self, filename: str, vacancies: list):
        self.filename = filename
        self.vacancies = vacancies

    def save_to_file(self):
        with open(self.filename, 'w', newline='', encoding='utf-8') as file:
            fieldnames = ['id_num', 'vacancy_name', 'company_name', 'description',
                          'salary_from', 'salary_to', 'salary_currency',
                          'area', 'url']
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.vacancies)

    def load_from_file(self):
        try:
            with open(self.filename, 'r', newline='', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                self.vacancies = [row for row in reader]
        except FileNotFoundError:
            self.vacancies = []

    def add_vacancy(self, vacancy):
        self.load_from_file()
        self.vacancies.append(vacancy)
        self.save_to_file()

    def get_vacancies(self, criteria):
        self.load_from_file()
        matching_vacancies = []
        for vacancy in self.vacancies:
            if criteria in vacancy['title'] or criteria in vacancy['description']:
                matching_vacancies.append(vacancy)
        return matching_vacancies

    def remove_vacancy(self, vacancy_id):
        self.load_from_file()
        self.vacancies = [vacancy for vacancy in self.vacancies if vacancy['id_num'] != vacancy_id]
        self.save_to_file()


class ExcelVacancyStorage(VacancyStorage):

    def __init__(self, filename: str, vacancies: list):
        """
        Initializes the ExcelVacancyStorage with the provided file name.

        Args:
            filename (str): The path to the Excel file used for storing vacancies.
        """
        self.filename = filename
        self.vacancies = vacancies

    def save_to_file(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['id', 'vacancy_name', 'company_name', 'description',
                      'salary_from', 'salary_to', 'salary_currency',
                      'area', 'url'])

        for vacancy in self.vacancies:
            sheet.append([vacancy['id_num'], vacancy['vacancy_name'], vacancy['company_name'],
                          vacancy['description'], vacancy['salary_from'], vacancy['salary_to'],
                          vacancy['salary_currency'], vacancy['area'], vacancy['url']])

        workbook.save(self.filename)

    def load_from_file(self):
        try:
            workbook = openpyxl.load_workbook(self.filename)
            sheet = workbook.active
            self.vacancies = [{'id_num': row[0].value, 'vacancy_name': row[1].value, 'company_name': row[2].value,
                               'description': row[3].value, 'salary_from': row[4].value, 'salary_to': row[5].value,
                               'salary_currency': row[6].value, 'area': row[7].value, 'url': row[8].value}
                              for row in sheet.iter_rows(min_row=2)]
        except FileNotFoundError:
            self.vacancies = []

    def add_vacancy(self, vacancy):
        self.load_from_file()
        self.vacancies.append(vacancy)
        self.save_to_file()

    def get_vacancies(self, criteria):
        self.load_from_file()
        matching_vacancies = []
        for vacancy in self.vacancies:
            if criteria in vacancy['title'] or criteria in vacancy['description']:
                matching_vacancies.append(vacancy)
        return matching_vacancies

    def remove_vacancy(self, vacancy_id):
        self.load_from_file()
        self.vacancies = [vacancy for vacancy in self.vacancies if vacancy['id_num'] != vacancy_id]
        self.save_to_file()


class TxtVacancyStorage(VacancyStorage):
    def __init__(self, filename: str, vacancies: list):
        """
        Инициализирует TxtVacancyStorage с указанным именем файла.

        Args:
            filename (str): Путь к текстовому файлу, используемому для хранения вакансий.
        """
        self.filename = filename
        self.vacancies = vacancies

    def save_to_file(self):
        with open(self.filename, 'w', encoding='utf-8') as file:
            for vacancy in self.vacancies:
                file.write(f"Id: {vacancy['id_num']}\n")
                file.write(f"Vacancy Name: {vacancy['vacancy_name']}\n")
                file.write(f"Company Name: {vacancy['company_name']}\n")
                file.write(f"Description: {vacancy['description']}\n")
                file.write(f"Salary From: {vacancy['salary_from']}\n")
                file.write(f"Salary To: {vacancy['salary_to']}\n")
                file.write(f"Salary Currency: {vacancy['salary_currency']}\n")
                file.write(f"Area: {vacancy['area']}\n")
                file.write(f"URL: {vacancy['url']}\n\n")

    def load_from_file(self):
        try:
            with open(self.filename, 'r', encoding='utf-8') as file:
                lines = file.readlines()
                vacancy = {}
                for line in lines:
                    if line.startswith("Id: "):
                        vacancy['id'] = line.strip()[4:]
                    elif line.startswith("Title: "):
                        vacancy['title'] = line.strip()[7:]
                    elif line.startswith("Description: "):
                        vacancy['description'] = line.strip()[12:]
                    elif line == "\n":
                        self.vacancies.append(vacancy)
                        vacancy = {}
        except FileNotFoundError:
            self.vacancies = []

    def add_vacancy(self, vacancy):
        self.load_from_file()
        self.vacancies.append(vacancy)
        self.save_to_file()

    def get_vacancies(self, criteria):
        self.load_from_file()
        matching_vacancies = []
        for vacancy in self.vacancies:
            if criteria in vacancy['title'] or criteria in vacancy['description']:
                matching_vacancies.append(vacancy)
        return matching_vacancies

    def remove_vacancy(self, vacancy_id):
        self.load_from_file()
        self.vacancies = [vacancy for vacancy in self.vacancies if vacancy['id_num'] != vacancy_id]
        self.save_to_file()


class DatabaseSaver:
    """
    Класс DatabaseSaver предназначен для сохранения данных о вакансиях в базу данных PostgreSQL.

    Attributes:
        connection (psycopg2.extensions.connection): Соединение с базой данных PostgreSQL.

    Methods:
        save_to_database(list_vacancies: list[dict]) -> None:
            Сохраняет список вакансий в базу данных. Перед сохранением выполняет очистку таблицы 'vacancies' и сброс
            счетчика первичного ключа.

    Example:
        # Создаем объект класса DatabaseSaver
        database_saver = DatabaseSaver()

        # Список вакансий для сохранения
        vacancies_data = [
            {
                'id_num': 1,
                'vacancy_name': 'Python Developer',
                'company_name': 'Example Company',
                'description': 'Job description...',
                'salary_from': 50000.0,
                'salary_to': 80000.0,
                'salary_currency': 'USD',
                'area': 'New York',
                'url': 'https://example.com/job/1'
            },
            # Другие вакансии...
        ]

        # Сохраняем вакансии в базу данных
        database_saver.save_to_database(vacancies_data)
    """

    def __init__(self):
        """
        Инициализирует объект класса DatabaseSaver и устанавливает соединение с базой данных PostgreSQL.
        """
        self.connection = psycopg2.connect(
            host="Localhost",
            database="postgres",
            user="postgres",
            password="admin@123",
        )

    def save_to_database(self, list_vacancies: list[dict]) -> None:
        """
        Сохраняет список вакансий в базу данных PostgreSQL.

        Args:
            list_vacancies (list[dict]): Список словарей, где каждый словарь представляет информацию о вакансии.
                Каждый словарь должен содержать ключи: 'id_num', 'vacancy_name', 'company_name', 'description',
                'salary_from', 'salary_to', 'salary_currency', 'area', 'url'.

        Returns:
            None
        """
        with self.connection as conn:
            with conn.cursor() as cursor:

                # Очищаем таблицу 'vacancies' и сбрасываем счетчик первичного ключа

                delete_query = "TRUNCATE TABLE vacancies RESTART IDENTITY;"
                cursor.execute(delete_query)
                cursor.execute(
                    """
                    CREATE TABLE vacancies (
                    id_num INT NOT NULL,
                    vacancy_name VARCHAR(255),
                    company_name VARCHAR(255),
                    description VARCHAR(255),
                    salary_from INT,
                    salary_to INT,
                    salary_currency VARCHAR(255),
                    area VARCHAR(255),
                    url VARCHAR(255)
                    )
                    """
                )

                # Сохраняем каждую вакансию из списка
                for item in list_vacancies:
                    id_num = item['id_num']
                    vacancy_name = item['vacancy_name']
                    company_name = item['company_name']
                    description = item['description']
                    salary_from = item['salary_from']
                    salary_to = item['salary_to']
                    salary_currency = item['salary_currency']
                    area = item['area']
                    url = item['url']

                    # SQL-запрос для вставки данных в таблицу 'vacancies'
                    query = (
                        "INSERT INTO vacancies"
                        "("
                        "id_num,"
                        "vacancy_name,"
                        "company_name,"
                        "description,"
                        "salary_from,"
                        "salary_to,"
                        "salary_currency,"
                        "area,"
                        "url"
                        ")"
                        "VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s);"
                    )
                    cursor.execute(query, (
                        id_num, vacancy_name, company_name, description, salary_from, salary_to, salary_currency, area,
                        url))
